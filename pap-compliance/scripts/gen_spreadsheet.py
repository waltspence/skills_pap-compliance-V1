#!/usr/bin/env python3
"""
pap-compliance/scripts/gen_spreadsheet.py

Generate PAP_Compliance_{date}.xlsx from search results and download log.
Summary sheet + per-day tabs. Color-coded statuses. Clickable PDF hyperlinks.

Usage:
    python scripts/gen_spreadsheet.py [--search /path/to/search_results.json] \
        [--downloads /path/to/download_log.json] [--output-dir /path/to/reports]
"""

import sys
import os
import json
import re
import argparse
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))
from utils import parse_avail_days, is_stale, REPORTS_DIR, OUTPUTS_DIR

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LINK_FONT = Font(color="0563C1", underline="single", size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))


def pick_portal_and_serial(patient):
    """Pick the best DOB-verified portal + serial, precedence AV > CO > RH."""
    for p in patient.get("av", []):
        if p.get("dob_verified"):
            return "AV", p.get("serial", "")
    for p in patient.get("co", []):
        if p.get("dob_verified"):
            return "CO", p.get("serial", "")
    for p in patient.get("rh", []):
        if p.get("dob_verified"):
            return "RH", p.get("serial", "")
    return "", ""


def get_status(patient, dl_lookup):
    """Return (status, fill, file_30, file_90)."""
    name = f"{patient['last']}, {patient['first']}"

    if name in dl_lookup:
        entry = dl_lookup[name]
        r30 = entry.get("report_30", {})
        r90 = entry.get("report_90")
        f30 = r30.get("file", "") if r30 and r30.get("status") == "OK" else ""
        f90 = r90.get("file", "") if r90 and r90.get("status") == "OK" else ""
        if f30 or f90:
            return "Downloaded", GREEN, f30, f90
        return "Download failed", RED, "", ""

    av_profiles = patient.get("av", [])
    co_matches = patient.get("co", [])
    rh_matches = patient.get("rh", [])

    if not av_profiles and not co_matches and not rh_matches:
        return "Not found", RED, "", ""

    if co_matches and not any(p.get("dob_verified") for p in av_profiles):
        return "CO only (manual pull)", YELLOW, "", ""

    verified = [p for p in av_profiles if p.get("dob_verified")]
    if not verified:
        for p in av_profiles:
            reason = p.get("skip_reason", "")
            if "stale" in reason:
                return f"Stale ({p.get('updated', '')})", YELLOW, "", ""
            if "zero_usage" in reason:
                return "0% usage", YELLOW, "", ""
            if "dob_mismatch" in reason:
                return "DOB mismatch (skipped)", RED, "", ""
        if is_stale(av_profiles[0].get("updated", "--")) if av_profiles else False:
            return "Bring device - data stale", YELLOW, "", ""
        return "Not downloadable", GRAY, "", ""

    best = verified[0]
    avail = parse_avail_days(best.get("available", "0"))
    if avail < 30:
        return f"Insufficient ({avail}d)", GRAY, "", ""
    if best.get("last30") == "0%":
        return "0% usage", YELLOW, "", ""

    return "Queued (not downloaded)", GRAY, "", ""


def set_link(cell, filename):
    """Make a cell a clickable relative hyperlink to a PDF in the same folder."""
    cell.value = filename
    cell.hyperlink = filename
    cell.font = LINK_FONT
    cell.border = THIN_BORDER


def bordered_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.border = THIN_BORDER
    return c


def write_headers(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def auto_width(ws, headers, last_row):
    for col in range(1, len(headers) + 1):
        max_len = len(headers[col - 1])
        for r in range(2, last_row):
            val = ws.cell(row=r, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 50)


def write_patient_row(ws, row, patient, dl_lookup, include_visit_date):
    name = f"{patient['last']}, {patient['first']}"
    portal, serial = pick_portal_and_serial(patient)
    status, fill, f30, f90 = get_status(patient, dl_lookup)

    col = 1
    bordered_cell(ws, row, col, name); col += 1
    bordered_cell(ws, row, col, patient.get("mrn", "")); col += 1
    bordered_cell(ws, row, col, patient["dob"]); col += 1
    if include_visit_date:
        bordered_cell(ws, row, col, patient.get("visit_date", "")); col += 1
    bordered_cell(ws, row, col, patient.get("provider", "")); col += 1
    bordered_cell(ws, row, col, portal); col += 1
    bordered_cell(ws, row, col, serial); col += 1
    sc = bordered_cell(ws, row, col, status); sc.fill = fill; col += 1
    if f30:
        set_link(ws.cell(row=row, column=col), f30)
    else:
        bordered_cell(ws, row, col, "")
    col += 1
    if f90:
        set_link(ws.cell(row=row, column=col), f90)
    else:
        bordered_cell(ws, row, col, "")
    col += 1
    bordered_cell(ws, row, col, patient.get("notes", patient.get("appt_notes", ""))[:80])


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--search", default="/home/claude/search_results.json")
    parser.add_argument("--downloads", default=os.path.join(REPORTS_DIR, "download_log.json"))
    parser.add_argument("--output-dir", default=REPORTS_DIR)
    parser.add_argument("--zip", action="store_true", help="Also create ZIP of all deliverables")
    args = parser.parse_args()

    with open(args.search) as f:
        search_results = json.load(f)

    dl_lookup = {}
    if os.path.exists(args.downloads):
        with open(args.downloads) as f:
            for entry in json.load(f):
                dl_lookup[entry["patient"]] = entry

    today = datetime.now().strftime("%m_%d_%Y")
    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    summary_headers = ["Patient", "MRN", "DOB", "Visit Date", "Provider",
                       "Portal", "Serial", "Status", "30-day", "90-day", "Notes"]
    write_headers(ws, summary_headers)

    row = 2
    for p in search_results:
        write_patient_row(ws, row, p, dl_lookup, include_visit_date=True)
        row += 1
    auto_width(ws, summary_headers, row)

    # Per-day sheets
    days = {}
    for p in search_results:
        vd = p.get("visit_date", "unknown")
        if vd not in days:
            days[vd] = []
        days[vd].append(p)

    day_headers = ["Patient", "MRN", "DOB", "Provider",
                   "Portal", "Serial", "Status", "30-day", "90-day", "Notes"]
    for day, pts in sorted(days.items()):
        safe = day.replace("/", "-") if day else "unknown"
        dws = wb.create_sheet(title=safe[:31])
        write_headers(dws, day_headers)

        drow = 2
        for p in pts:
            write_patient_row(dws, drow, p, dl_lookup, include_visit_date=False)
            drow += 1
        auto_width(dws, day_headers, drow)

    xlsx_path = os.path.join(args.output_dir, f"PAP_Compliance_{today}.xlsx")
    wb.save(xlsx_path)
    print(f"Spreadsheet: {xlsx_path}")

    # Run log
    log_lines = [
        f"PAP Compliance Run Log — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "=" * 60, "",
    ]

    statuses = {}
    for p in search_results:
        s, _, _, _ = get_status(p, dl_lookup)
        statuses[s] = statuses.get(s, 0) + 1
    log_lines.append("Status summary:")
    for s, c in sorted(statuses.items(), key=lambda x: -x[1]):
        log_lines.append(f"  {s}: {c}")

    log_path = os.path.join(args.output_dir, f"PAP_Compliance_Log_{today}.txt")
    with open(log_path, "w") as f:
        f.write("\n".join(log_lines))
    print(f"Log: {log_path}")

    if args.zip:
        zip_path = os.path.join(OUTPUTS_DIR, f"PAP_Compliance_{today}.zip")
        os.makedirs(OUTPUTS_DIR, exist_ok=True)
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for fname in sorted(os.listdir(args.output_dir)):
                fpath = os.path.join(args.output_dir, fname)
                if os.path.isfile(fpath):
                    zf.write(fpath, fname)
        print(f"ZIP: {zip_path}")

    print(f"\nStatus counts:")
    for s, c in sorted(statuses.items(), key=lambda x: -x[1]):
        print(f"  {s}: {c}")


if __name__ == "__main__":
    main()
